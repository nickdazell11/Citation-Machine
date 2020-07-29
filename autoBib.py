r'''
TO ENSURE THE PROGRAM RUNS SMOOTHLY...

Make sure not to rename the sheets in the excel file or create any new sheets.
Don't add headers, titles, or any other formatting.
List the sources in the first column, 1 in each row, with no empty rows until the end of the source list.
Start by installing selenium and pyperclip by typing the following command into the terminal:
pip install selenium
pip install pyperclip
Download firefox web browser and geckodriver (the Gecko Driver folder needs to be added to your PATH)
Might need admin control to allow webDriver the first time
'''

import openpyxl
from openpyxl import Workbook
from selenium import webdriver
import time
import pyperclip
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException, TimeoutException, StaleElementReferenceException


# sources should be of the format 'filename.xlsx' i.e. 'sourceList.xlsx'
def citMachine(sources): # i.e. citMachine('sourceList.xlsx')

    # Load the input excel file
    wb = openpyxl.load_workbook(sources)

    # Name the first sheet Source List
    sheet1 = wb.worksheets[0]
    sheet1.title = 'Source List'

    # Make sheet 2 if needed, name the second sheet Citations
    num_sheets = len(wb.worksheets)
    if num_sheets < 2:
        wb.create_sheet('Citations')
    sheet2 = wb.worksheets[1]
    if num_sheets >= 2:
        sheet2.title = 'Citations'

    # Resize the columns to make URL and citation more visible
    sheet2.column_dimensions['A'].width = 50
    sheet2.column_dimensions['B'].width = 210
    

    # Copy each entry in the first sheet to the second sheet and get the citation for each source from citationmachine.net
    for i in range(1,100):
        if sheet1.cell(row = i, column = 1).value != None:
            if sheet2.cell(row = i, column = 2).value == None:   

                # Copy the sources from sheet1 to sheet2
                sheet2.cell(row = i, column = 1).value = sheet1.cell(row = i, column = 1).value

                # open the firefox browser and navigate to citationmachine.net
                driver = webdriver.Firefox()
                driver.get('http://www.citationmachine.net')

                # click on 'Create Citations'
                wait = WebDriverWait(driver, 10)
                wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/main/div[2]/div/div[1]/a')))
                createCitations = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/main/div[2]/div/div[1]/a')
                createCitations.click()

                # Click on 'Website'
                wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/section/div/div[2]/div/button[1]')))
                website = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/section/div/div[2]/div/button[1]')
                website.click()

                # Type in the source URL to the search bar and hit enter
                wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/section/div/div[2]/div/input')))
                sourceURL = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/section/div/div[2]/div/input')
                sourceURL.send_keys(sheet1.cell(row = i, column = 1).value)
                sourceURL.send_keys(Keys.RETURN)

                # Click on 'Cite'
                wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, 'button.styled__ResultButton-jhqr36-12')))
                cite = driver.find_element_by_css_selector("button.styled__ResultButton-jhqr36-12")
                cite.click()

                # Click 'Continue'
                wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div[2]/div/div/div/div/a')))
                continueButton = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div[2]/div/div/div/div/a')
                driver.execute_script("arguments[0].click();", continueButton)
                try:
                    continueButton.click()
                    break
                except(ElementClickInterceptedException):
                    pass
                
                # Click on 'Complete Citation'
                time.sleep(5)
                driver.refresh()
                wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, 'button.styled__SubmitButton-t0wj53-1')))
                completeCitation = driver.find_element_by_css_selector('button.styled__SubmitButton-t0wj53-1')
                driver.execute_script("arguments[0].click();", completeCitation)
                '''
                try:
                    completeCitation.click()
                    break
                except(ElementClickInterceptedException, TimeoutException):
                    pass
                '''
                # Try to watch an advertisement in case it pops up
                try:
                    wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[11]/div/div/div/div/a[1]')))
                    driver.find_element_by_xpath('/html/body/div[11]/div/div/div/div/a[1]').click()
                    time.sleep(32)
                    break
                except(NoSuchElementException, TimeoutException):
                    pass

                # Click on Change Style
                wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, 'div.styled__ButtonIcon-sc-1xmkaq-3')))
                changeStyle = driver.find_element_by_css_selector('div.styled__ButtonIcon-sc-1xmkaq-3')
                # changeStyle.click()
                driver.execute_script("arguments[0].click();", changeStyle)

                # Send Chicago Keys
                wait.until(expected_conditions.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div/div/div[1]/input')))
                chicago = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div[1]/div/div[2]/div[2]/div/div/div[1]/input')
                chicago.send_keys('Chicago')
                time.sleep(2)
                chicago.send_keys(Keys.DOWN)
                chicago.send_keys(Keys.RETURN)

                # Click 'Copy All'
                try:
                    wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, 'button.styled__Button-sc-1xmkaq-1')))
                    option1 = driver.find_element_by_css_selector('button.styled__Button-sc-1xmkaq-1')
                    driver.execute_script("arguments[0].click();", option1)
                except:
                    driver.refresh()
                    print('driver refreshed')
                    wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, 'button.styled__Button-sc-1xmkaq-1')))
                    option1 = driver.find_element_by_css_selector('button.styled__Button-sc-1xmkaq-1')
                    driver.execute_script("arguments[0].click();", option1)

                # Close the driver
                driver.close()

                # Paste the copied text from the clipboard into excel
                print('Citation #' + str(i) + ' completed')
                sheet2.cell(row = i, column = 2).value = pyperclip.paste()
                
                # Save the updated excel file (could be taken out of loop to reduce runtime?)
                wb.save(sources)

    # Save and close the excel file
    wb.save(sources)
    wb.close()